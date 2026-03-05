"""
Integration Tests for GlassOrchestrator — System handshakes between components.

IT-1: Gmail Connection & Search  (requires live credentials — skipped by default)
IT-2: Worker Handoff (File Bridge)
IT-3: Merge Reconciliation
IT-4: Spreadsheet Persistence
"""

import csv
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from GlassOrchestrator import (
    COLUMNS,
    CSV_PATH,
    DATA_DIR,
    IMAP_SERVER,
    SHEET_NAME,
    TARGET_SENDER,
    phase2_parse,
    phase4_merge,
    phase5_persist,
)


# ─── IT-1: Gmail Connection & Search ─────────────────────────────────────────


@pytest.mark.skipif(
    not os.getenv("GLASS_EMAIL_ACCOUNT") or not os.getenv("GLASS_EMAIL_PASSWORD"),
    reason="Gmail credentials not configured (set GLASS_EMAIL_ACCOUNT / GLASS_EMAIL_PASSWORD)",
)
class TestIT1_GmailConnection:
    """Verify IMAP authentication and UNSEEN message search."""

    def test_imap_auth_and_search(self):
        import imaplib

        account = os.getenv("GLASS_EMAIL_ACCOUNT")
        password = os.getenv("GLASS_EMAIL_PASSWORD")

        mail = imaplib.IMAP4_SSL(IMAP_SERVER)
        try:
            mail.login(account, password)
            mail.select("inbox")
            status, msg_ids = mail.search(None, f'(FROM "{TARGET_SENDER}")')
            assert status == "OK"
            # msg_ids[0] may be empty (no messages) — that's still OK
            count = len(msg_ids[0].split()) if msg_ids[0] else 0
            assert isinstance(count, int)
        finally:
            mail.logout()


# ─── IT-2: Worker Handoff (File Bridge) ──────────────────────────────────────


class TestIT2_WorkerHandoff:
    """Verify CSV formatting and subprocess launch capability."""

    def test_csv_written_correctly(self, tmp_path):
        """Write MVAs to a CSV and validate structure."""
        csv_path = tmp_path / "GlassDataParser.csv"
        mva_list = ["59340120", "59340121", "59340122"]

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["MVA"])
            for mva in mva_list:
                writer.writerow([mva])

        # Validate
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            assert "MVA" in reader.fieldnames
            rows = list(reader)
            assert len(rows) == 3
            assert rows[0]["MVA"] == "59340120"
            assert rows[2]["MVA"] == "59340122"

    def test_subprocess_can_launch_python(self):
        """Verify subprocess.check_call can invoke Python successfully."""
        # Run a trivial Python command to prove the subprocess mechanism works
        subprocess.check_call(
            [sys.executable, "-c", "print('worker ok')"],
        )

    def test_subprocess_detects_failure(self):
        """Verify subprocess.CalledProcessError is raised on non-zero exit."""
        with pytest.raises(subprocess.CalledProcessError):
            subprocess.check_call(
                [sys.executable, "-c", "import sys; sys.exit(1)"],
            )


# ─── IT-3: Merge Reconciliation ──────────────────────────────────────────────


class TestIT3_MergeReconciliation:
    """Verify left-join logic: missing MVAs produce VIN='N/A', no rows dropped."""

    def test_left_join_with_missing_mva(self, tmp_path, monkeypatch):
        """One MVA present in results, one missing → VIN=N/A for missing."""
        # Build manifest from Phase 2
        descriptions = ["59340120", "59340121"]
        manifest, _ = phase2_parse(descriptions, datetime(2026, 3, 5))

        # Create a mock GlassResults.txt with only ONE of the two MVAs
        results_file = tmp_path / "GlassResults.txt"
        results_file.write_text("MVA,VIN,Desc\n59340120,1HGCM82633A004352,Windshield\n")

        # Patch the RESULTS_PATH used by phase4_merge
        monkeypatch.setattr("GlassOrchestrator.RESULTS_PATH", results_file)

        df = phase4_merge(manifest)

        assert len(df) == 2  # Both MVAs must be present (no drop)
        row_120 = df[df["MVA"] == "59340120"].iloc[0]
        row_121 = df[df["MVA"] == "59340121"].iloc[0]

        assert row_120["VIN"] == "1HGCM82633A004352"
        assert row_121["VIN"] == "N/A"  # Missing from scraper → N/A

    def test_all_mvas_missing_from_results(self, tmp_path, monkeypatch):
        """No scraper results at all → all VINs become N/A."""
        descriptions = ["59340120", "59340121"]
        manifest, _ = phase2_parse(descriptions, datetime(2026, 3, 5))

        results_file = tmp_path / "GlassResults.txt"
        results_file.write_text("MVA,VIN,Desc\n")  # header only

        monkeypatch.setattr("GlassOrchestrator.RESULTS_PATH", results_file)

        df = phase4_merge(manifest)

        assert len(df) == 2
        assert (df["VIN"] == "N/A").all()

    def test_results_file_missing(self, tmp_path, monkeypatch):
        """No GlassResults.txt file → degrade gracefully, all VINs = N/A."""
        descriptions = ["59340120"]
        manifest, _ = phase2_parse(descriptions, datetime(2026, 3, 5))

        monkeypatch.setattr("GlassOrchestrator.RESULTS_PATH", tmp_path / "nonexistent.txt")

        df = phase4_merge(manifest)
        assert len(df) == 1
        assert df.iloc[0]["VIN"] == "N/A"


# ─── IT-4: Spreadsheet Persistence ───────────────────────────────────────────


class TestIT4_SpreadsheetPersistence:
    """Verify data is appended to the correct sheet/columns without overwriting."""

    def _make_test_df(self, mvas, date="2026-03-05"):
        rows = []
        for mva in mvas:
            rows.append({
                "Arrival Date": date,
                "MVA": mva,
                "VIN": "1HGCM82633A004352",
                "Make": "Windshield",
                "Location": "APO",
                "Damage Type": "Replacement",
                "Claim#": "Pending",
                "WorkItem": "verified",
            })
        return pd.DataFrame(rows, columns=COLUMNS)

    def test_creates_new_workbook(self, tmp_path, monkeypatch):
        """First write creates file with headers + data."""
        test_log = tmp_path / "TestLog.xlsx"
        monkeypatch.setattr("GlassOrchestrator.MASTER_LOG", test_log)

        df = self._make_test_df(["59340120", "59340121"])
        new_rows = phase5_persist(df)

        assert test_log.exists()
        assert len(new_rows) == 2

        wb = load_workbook(str(test_log))
        ws = wb[SHEET_NAME]
        headers = [cell.value for cell in ws[1]]
        assert headers == COLUMNS
        assert ws.cell(row=2, column=2).value == "59340120"
        assert ws.cell(row=3, column=2).value == "59340121"
        wb.close()

    def test_appends_without_overwriting(self, tmp_path, monkeypatch):
        """Second write appends new rows; original rows untouched."""
        test_log = tmp_path / "TestLog.xlsx"
        monkeypatch.setattr("GlassOrchestrator.MASTER_LOG", test_log)

        # First batch
        df1 = self._make_test_df(["59340120"])
        phase5_persist(df1)

        # Second batch (different MVA)
        df2 = self._make_test_df(["59340121"])
        new_rows = phase5_persist(df2)

        assert len(new_rows) == 1

        wb = load_workbook(str(test_log))
        ws = wb[SHEET_NAME]
        assert ws.max_row == 3  # header + 2 data rows
        assert ws.cell(row=2, column=2).value == "59340120"  # original intact
        assert ws.cell(row=3, column=2).value == "59340121"  # appended
        wb.close()

    def test_idempotency_prevents_duplicate(self, tmp_path, monkeypatch):
        """Same MVA+Date written twice → only one row stored."""
        test_log = tmp_path / "TestLog.xlsx"
        monkeypatch.setattr("GlassOrchestrator.MASTER_LOG", test_log)

        df = self._make_test_df(["59340120"])
        phase5_persist(df)
        new_rows = phase5_persist(df.copy())

        assert len(new_rows) == 0

        wb = load_workbook(str(test_log))
        ws = wb[SHEET_NAME]
        assert ws.max_row == 2  # header + 1 data row (no duplicate)
        wb.close()

    def test_correct_columns_written(self, tmp_path, monkeypatch):
        """Verify all 8 columns match the expected schema."""
        test_log = tmp_path / "TestLog.xlsx"
        monkeypatch.setattr("GlassOrchestrator.MASTER_LOG", test_log)

        df = self._make_test_df(["59340120"])
        phase5_persist(df)

        wb = load_workbook(str(test_log))
        ws = wb[SHEET_NAME]
        row_data = [cell.value for cell in ws[2]]
        assert row_data == ["2026-03-05", "59340120", "1HGCM82633A004352",
                            "Windshield", "APO", "Replacement", "Pending", "verified"]
        wb.close()
