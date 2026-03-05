"""
GlassOrchestrator.py — 6-Phase Vehicle Glass Procurement Pipeline

Phases:
  1. Input       – Fetch scan data from Gmail (export@orcascan.com)
  2. Parsing     – Regex triage, build session manifest
  3. Worker      – Write CSV, invoke GlassDataParser.py subprocess
  4. Data Merge  – Left-join manifest with scraper results
  5. Persistence – Append to MasterLog.xlsx (idempotent)
  6. Notification – HTML email for Replacement items
"""

import csv
import email
import imaplib
import logging
import os
import re
import smtplib
import subprocess
import sys
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

# ─── Configuration ────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
CSV_PATH = DATA_DIR / "GlassDataParser.csv"
RESULTS_PATH = BASE_DIR / "GlassResults.txt"
MASTER_LOG = BASE_DIR / "MasterLog.xlsx"
SHEET_NAME = "GlassClaims"
WORKER_SCRIPT = BASE_DIR / "CGI" / "src" / "GlassDataParser.py"

# Gmail credentials — set via environment variables
IMAP_SERVER = "imap.gmail.com"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_ACCOUNT = os.getenv("GLASS_EMAIL_ACCOUNT", "")
EMAIL_PASSWORD = os.getenv("GLASS_EMAIL_PASSWORD", "")  # App password recommended
SENDER_ADDRESS = os.getenv("GLASS_SENDER", "")
NOTIFY_RECIPIENTS = os.getenv("GLASS_NOTIFY_RECIPIENTS", "").split(",")

TARGET_SENDER = "export@orcascan.com"
MVA_PATTERN = re.compile(r"^(\d{8})([rc]*)$")

LOCATION = "APO"

COLUMNS = [
    "Date",
    "MVA",
    "VIN",
    "Description",
    "Location",
    "WorkType",
    "ClaimStatus",
]

# ─── Logging ──────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(BASE_DIR / "GlassOrchestrator.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("GlassOrchestrator")

# ─── Phase 1 — Input (Inbound Acquisition) ───────────────────────────────────


def phase1_input() -> tuple[list[str], datetime]:
    """
    Connect to Gmail via IMAP, fetch the latest UNSEEN email from the
    target sender, and extract:
      - A list of raw 'Description' strings (one per line in the body/CSV)
      - The Date header parsed as a datetime object
    """
    log.info("PHASE 1 — Input: Connecting to Gmail IMAP …")

    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    try:
        mail.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
        mail.select("inbox")

        search_criteria = f'(UNSEEN FROM "{TARGET_SENDER}")'
        status, msg_ids = mail.search(None, search_criteria)
        if status != "OK" or not msg_ids[0]:
            log.warning("Phase 1: No unseen messages from %s", TARGET_SENDER)
            return [], datetime.now()

        ids = msg_ids[0].split()
        log.info("Phase 1: Found %d unseen message(s)", len(ids))

        # Process only the latest message
        latest_id = ids[-1]
        status, msg_data = mail.fetch(latest_id, "(RFC822)")
        if status != "OK":
            raise RuntimeError(f"Failed to fetch message id {latest_id}")

        raw_email = msg_data[0][1]
        msg = email.message_from_bytes(raw_email)

        # Parse date
        date_str = msg.get("Date", "")
        email_date = email.utils.parsedate_to_datetime(date_str) if date_str else datetime.now()
        log.info("Phase 1: Email date = %s", email_date.isoformat())

        # Extract body
        body = _extract_body(msg)
        # Try HTML table extraction first, then fall back to CSV/plain-text
        if "<table" in body.lower():
            descriptions = _parse_html_descriptions(body)
        else:
            descriptions = _parse_descriptions(body)
        log.info("Phase 1: Extracted %d description lines", len(descriptions))

        # Mark as Seen (IMAP already marks on FETCH unless PEEK is used)
        return descriptions, email_date

    finally:
        try:
            mail.logout()
        except Exception:
            pass


def _extract_body(msg: email.message.Message) -> str:
    """Walk a MIME message and return the best body for parsing.

    Orca Scan emails contain an HTML table with structured data and a
    plain-text CSV attachment.  We prefer the HTML when it contains a
    <table> because the CSV embeds newlines inside quoted fields which
    break simple line-by-line parsing.
    """
    plain = ""
    html = ""
    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            payload = part.get_payload(decode=True)
            if not payload:
                continue
            decoded = payload.decode("utf-8", errors="replace")
            if content_type == "text/plain" and not plain:
                plain = decoded
            elif content_type == "text/html" and not html:
                html = decoded
    else:
        content_type = msg.get_content_type()
        payload = msg.get_payload(decode=True)
        if payload:
            decoded = payload.decode("utf-8", errors="replace")
            if content_type == "text/html":
                html = decoded
            else:
                plain = decoded
    # Prefer HTML when it contains a data table (Orca Scan format)
    if html and "<table" in html.lower():
        return html
    return plain or html


def _parse_descriptions(body: str) -> list[str]:
    """
    Parse the email body as CSV or line-delimited text and return
    the 'Description' column values.
    """
    lines = body.strip().splitlines()
    if not lines:
        return []

    # Try CSV with a 'Description' header first
    reader = csv.DictReader(lines)
    if reader.fieldnames and "Description" in reader.fieldnames:
        return [row["Description"].strip() for row in reader if row.get("Description", "").strip()]

    # Fallback: treat each non-empty line as a description
    return [line.strip() for line in lines if line.strip()]


def _parse_html_descriptions(html: str) -> list[str]:
    """
    Extract 'Description' column values from an HTML table (Orca Scan email).
    Uses BeautifulSoup if available, otherwise falls back to regex.
    """
    if HAS_BS4:
        soup = BeautifulSoup(html, "html.parser")
        # Prefer the data table (id="rowData") over layout tables
        table = soup.find("table", id="rowData") or soup.find("table")
        if not table:
            return []
        rows = table.find_all("tr")
        if not rows:
            return []
        # Find Description column index from header row
        headers = [th.get_text(strip=True) for th in rows[0].find_all(["th", "td"])]
        if "Description" not in headers:
            return []
        desc_idx = headers.index("Description")
        descriptions = []
        for row in rows[1:]:
            cells = row.find_all("td")
            if len(cells) > desc_idx:
                # Orca Scan packs multiple MVAs into one cell separated by newlines
                raw = cells[desc_idx].get_text(separator="\n", strip=False)
                for line in raw.splitlines():
                    line = line.strip()
                    if line:
                        descriptions.append(line)
        return descriptions
    else:
        # Regex fallback: extract td contents from rows
        # Try to isolate the data table (id="rowData") first
        table_match = re.search(
            r'<table[^>]*id=["\']rowData["\'][^>]*>(.*?)</table>',
            html, re.DOTALL | re.IGNORECASE,
        )
        search_html = table_match.group(1) if table_match else html
        # Find header row to locate Description column
        header_match = re.search(
            r"<tr[^>]*>(.*?)</tr>", search_html, re.DOTALL | re.IGNORECASE
        )
        if not header_match:
            return []
        header_cells = re.findall(
            r"<t[hd][^>]*>(.*?)</t[hd]>", header_match.group(1), re.DOTALL | re.IGNORECASE
        )
        header_cells = [re.sub(r"<[^>]+>", "", c).strip() for c in header_cells]
        if "Description" not in header_cells:
            return []
        desc_idx = header_cells.index("Description")
        # Extract data rows
        all_rows = re.findall(
            r"<tr[^>]*>(.*?)</tr>", search_html, re.DOTALL | re.IGNORECASE
        )
        descriptions = []
        for row_html in all_rows[1:]:  # skip header
            cells = re.findall(
                r"<td[^>]*>(.*?)</td>", row_html, re.DOTALL | re.IGNORECASE
            )
            cells = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]
            if len(cells) > desc_idx and cells[desc_idx]:
                # Split multi-line cell into individual MVAs
                for line in cells[desc_idx].splitlines():
                    line = line.strip()
                    if line:
                        descriptions.append(line)
        return descriptions


# ─── Phase 2 — Parsing (Triage & Normalization) ──────────────────────────────


def phase2_parse(descriptions: list[str], email_date: datetime) -> tuple[dict, list[str]]:
    """
    Apply regex to each description string and build a session manifest.

    Returns:
        manifest: dict keyed by MVA → {WorkType, ClaimStatus, Description, Date, Location}
        mva_list: list of clean 8-digit MVA strings for the worker
    """
    log.info("PHASE 2 — Parsing: Processing %d descriptions …", len(descriptions))

    manifest: dict[str, dict] = {}
    mva_list: list[str] = []
    date_str = email_date.strftime("%Y-%m-%d")

    for desc in descriptions:
        match = MVA_PATTERN.match(desc.strip())
        if not match:
            log.warning("Phase 2: Malformed entry skipped — '%s'", desc)
            continue

        mva = match.group(1)
        suffixes = match.group(2)

        work_type = "Repair" if "r" in suffixes else "Replacement"
        claim_status = "Claim Generated" if "c" in suffixes else "Pending"

        manifest[mva] = {
            "Date": date_str,
            "MVA": mva,
            "VIN": "",  # Populated in Phase 4
            "Description": desc.strip(),
            "Location": LOCATION,
            "WorkType": work_type,
            "ClaimStatus": claim_status,
        }
        mva_list.append(mva)

    log.info("Phase 2: Manifest built — %d valid MVAs, %d malformed",
             len(manifest), len(descriptions) - len(manifest))
    return manifest, mva_list


# ─── Phase 3 — Worker Processing (Enrichment) ────────────────────────────────


def phase3_worker(mva_list: list[str]) -> None:
    """
    Write clean MVAs to the CSV interface file, then invoke the
    external GlassDataParser.py worker as a subprocess.

    Raises subprocess.CalledProcessError on worker failure.
    """
    log.info("PHASE 3 — Worker: Writing %d MVAs to %s …", len(mva_list), CSV_PATH)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["MVA"])
        for mva in mva_list:
            writer.writerow([mva])

    log.info("Phase 3: Invoking worker subprocess — %s", WORKER_SCRIPT)
    subprocess.check_call(
        [sys.executable, str(WORKER_SCRIPT)],
        cwd=str(BASE_DIR),
    )
    log.info("Phase 3: Worker completed successfully")


def validate_results_freshness(results_path: Path, max_age_seconds: int = 300) -> None:
    """
    Verify that the results file was recently modified (by the current worker run).
    Raises RuntimeError if the file is stale or missing.
    """
    if not results_path.exists():
        raise RuntimeError(f"Results file not found: {results_path}")
    mtime = datetime.fromtimestamp(results_path.stat().st_mtime)
    age = datetime.now() - mtime
    if age > timedelta(seconds=max_age_seconds):
        raise RuntimeError(
            f"Stale results file: {results_path} was last modified "
            f"{age} ago (max allowed: {max_age_seconds}s)"
        )


# ─── Phase 4 — Data Merge (Reconciliation) ───────────────────────────────────


def phase4_merge(manifest: dict) -> pd.DataFrame:
    """
    Left-join the session manifest (Phase 2) with scraper output
    (GlassResults.txt from Phase 3).

    Any scanned MVA without a scraper match keeps VIN='N/A'.
    """
    log.info("PHASE 4 — Merge: Reconciling manifest with scraper results …")

    # Build manifest DataFrame
    df_manifest = pd.DataFrame(list(manifest.values()))

    # Read scraper results
    if RESULTS_PATH.exists():
        df_results = pd.read_csv(
            RESULTS_PATH,
            sep=",",
            dtype=str,
            encoding="utf-8",
        )
        # Normalize column names
        df_results.columns = [c.strip() for c in df_results.columns]
        # Keep only MVA and VIN (and optionally Desc from scraper)
        result_cols = [c for c in ["MVA", "VIN", "Desc"] if c in df_results.columns]
        df_results = df_results[result_cols]
    else:
        log.warning("Phase 4: %s not found — all VINs will be N/A", RESULTS_PATH)
        df_results = pd.DataFrame(columns=["MVA", "VIN"])

    # Left join
    df_merged = df_manifest.merge(
        df_results[["MVA", "VIN"]].rename(columns={"VIN": "VIN_scraped"}),
        on="MVA",
        how="left",
    )

    # Populate VIN: scraped value if available, else 'N/A'
    df_merged["VIN"] = df_merged["VIN_scraped"].fillna("N/A")
    df_merged.drop(columns=["VIN_scraped"], inplace=True)

    # Ensure column order
    df_merged = df_merged[COLUMNS]

    n_missing = (df_merged["VIN"] == "N/A").sum()
    log.info("Phase 4: Merge complete — %d rows, %d missing VINs", len(df_merged), n_missing)
    return df_merged


# ─── Phase 5 — Persistence (System of Record) ────────────────────────────────


def phase5_persist(df: pd.DataFrame) -> pd.DataFrame:
    """
    Append merged data to MasterLog.xlsx on the 'GlassClaims' sheet.

    Idempotency: Composite key (MVA + Date) is checked against the last
    100 rows of the existing sheet. Duplicate rows are silently skipped.

    Returns the DataFrame of actually-new rows written.
    """
    log.info("PHASE 5 — Persistence: Appending to %s [%s] …", MASTER_LOG, SHEET_NAME)

    # Determine which rows are truly new
    existing_keys = _load_existing_keys(MASTER_LOG, SHEET_NAME)


    # Filter out duplicates
    df["_key"] = df["MVA"] + "|" + df["Date"]
    new_rows = df[~df["_key"].isin(existing_keys)].drop(columns=["_key"]).copy()
    df.drop(columns=["_key"], inplace=True)

    if new_rows.empty:
        log.info("Phase 5: No new rows to write (all duplicates)")
        return new_rows

    # Append to workbook
    if MASTER_LOG.exists():
        wb = load_workbook(str(MASTER_LOG))
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(COLUMNS)
        else:
            ws = wb[SHEET_NAME]
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(COLUMNS)

    for _, row in new_rows.iterrows():
        ws.append([row[col] for col in COLUMNS])

    wb.save(str(MASTER_LOG))
    wb.close()

    log.info("Phase 5: Wrote %d new rows to %s", len(new_rows), MASTER_LOG)
    return new_rows


def _load_existing_keys(log_path: Path, sheet_name: str) -> set[str]:
    """
    Read the last 100 data rows from the given Excel workbook/sheet and
    return a set of 'MVA|Date' composite keys for idempotency checking.
    """
    existing_keys: set[str] = set()
    if not log_path.exists():
        return existing_keys
    try:
        wb = load_workbook(str(log_path))
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row or 1
            start_row = max(2, max_row - 99)
            headers = [cell.value for cell in ws[1]]
            mva_idx = headers.index("MVA") if "MVA" in headers else None
            date_idx = headers.index("Date") if "Date" in headers else None
            if mva_idx is not None and date_idx is not None:
                for row in ws.iter_rows(min_row=start_row, max_row=max_row, values_only=True):
                    key = f"{row[mva_idx]}|{row[date_idx]}"
                    existing_keys.add(key)
        wb.close()
    except Exception as exc:
        log.warning("Could not read existing workbook — %s", exc)
    return existing_keys


def is_duplicate(mva: str, date: str, existing_keys: set[str]) -> bool:
    """Return True if the MVA+Date composite key already exists."""
    return f"{mva}|{date}" in existing_keys


# ─── Phase 6 — Notification (Distribution) ───────────────────────────────────


def phase6_notify(df: pd.DataFrame) -> None:
    """
    Filter for Replacement items, build an HTML email with a styled table,
    and send it. Rows with VIN='N/A' are highlighted red to flag the
    ordering team for manual action.
    """
    log.info("PHASE 6 — Notification: Building replacement alert …")

    replacements = df[df["WorkType"] == "Replacement"]
    if replacements.empty:
        log.info("Phase 6: No Replacement items — skipping notification")
        return

    html = _build_html_table(replacements)
    subject = f"Glass Replacement Order — {replacements.iloc[0]['Date']} ({len(replacements)} items)"

    _send_email(subject, html)
    log.info("Phase 6: Notification sent to %s", NOTIFY_RECIPIENTS)


def _build_html_table(df: pd.DataFrame) -> str:
    """Render a DataFrame to an HTML table with red highlighting for N/A VINs."""
    has_missing = (df["VIN"] == "N/A").any()

    rows_html = ""
    for _, row in df.iterrows():
        if row["VIN"] == "N/A":
            style = ' style="background-color:#ff4444; color:#ffffff; font-weight:bold;"'
            vin_cell = (
                '<td style="background-color:#ff0000; color:#ffffff; font-weight:bold;">'
                '⚠ N/A — ACTION REQUIRED</td>'
            )
        else:
            style = ""
            vin_cell = f"<td>{row['VIN']}</td>"

        rows_html += f"""<tr{style}>
            <td>{row['Date']}</td>
            <td>{row['MVA']}</td>
            {vin_cell}
            <td>{row['Description']}</td>
            <td>{row['Location']}</td>
            <td>{row['WorkType']}</td>
            <td>{row['ClaimStatus']}</td>
        </tr>\n"""

    alert_banner = ""
    if has_missing:
        alert_banner = """
        <div style="background-color:#ff4444; color:#ffffff; padding:12px;
                    font-size:16px; font-weight:bold; margin-bottom:16px;
                    border-radius:4px; text-align:center;">
            ⚠ ATTENTION: One or more VINs could not be retrieved.
            Rows highlighted in RED require manual lookup before ordering.
        </div>
        """

    return f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; padding: 16px; }}
            table {{ border-collapse: collapse; width: 100%; }}
            th {{ background-color: #333; color: #fff; padding: 10px; text-align: left; }}
            td {{ border: 1px solid #ddd; padding: 8px; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
        </style>
    </head>
    <body>
        <h2>Glass Replacement Order Summary</h2>
        {alert_banner}
        <table>
            <thead>
                <tr>
                    <th>Date</th><th>MVA</th><th>VIN</th>
                    <th>Description</th><th>Location</th>
                    <th>Work Type</th><th>Claim Status</th>
                </tr>
            </thead>
            <tbody>
                {rows_html}
            </tbody>
        </table>
        <p style="margin-top:16px; font-size:12px; color:#888;">
            Generated by GlassOrchestrator &mdash; {datetime.now().strftime('%Y-%m-%d %H:%M')}
        </p>
    </body>
    </html>
    """


def _send_email(subject: str, html_body: str) -> None:
    """Send an HTML email via Gmail SMTP."""
    if not SENDER_ADDRESS or not NOTIFY_RECIPIENTS[0]:
        log.warning("Phase 6: Email credentials not configured — printing HTML to log")
        log.info("Subject: %s", subject)
        return

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = SENDER_ADDRESS
    msg["To"] = ", ".join(NOTIFY_RECIPIENTS)
    msg.attach(MIMEText(html_body, "html"))

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.ehlo()
        server.starttls()
        server.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
        server.sendmail(SENDER_ADDRESS, NOTIFY_RECIPIENTS, msg.as_string())


# ─── Pipeline Orchestrator ────────────────────────────────────────────────────


def run_pipeline() -> None:
    """Execute the full 6-phase pipeline with phase-level error handling."""
    log.info("=" * 60)
    log.info("GlassOrchestrator pipeline starting")
    log.info("=" * 60)

    # ── Phase 1: Input ────────────────────────────────────────
    try:
        descriptions, email_date = phase1_input()
    except Exception as exc:
        log.error("PHASE 1 FAILED — %s", exc, exc_info=True)
        return

    if not descriptions:
        log.info("Pipeline complete — no descriptions to process")
        return

    # ── Phase 2: Parsing ──────────────────────────────────────
    try:
        manifest, mva_list = phase2_parse(descriptions, email_date)
    except Exception as exc:
        log.error("PHASE 2 FAILED — %s", exc, exc_info=True)
        return

    if not manifest:
        log.info("Pipeline complete — no valid MVAs after parsing")
        return

    # ── Phase 3: Worker ───────────────────────────────────────
    try:
        phase3_worker(mva_list)
    except subprocess.CalledProcessError as exc:
        log.error("PHASE 3 FAILED — Worker returned non-zero exit code %d. "
                   "Pipeline ABORTED. No data will be persisted.", exc.returncode)
        return
    except Exception as exc:
        log.error("PHASE 3 FAILED — %s. Pipeline ABORTED.", exc, exc_info=True)
        return

    # ── Phase 3b: Validate freshness of results ───────────────
    try:
        validate_results_freshness(RESULTS_PATH)
    except RuntimeError as exc:
        log.error("PHASE 3 VALIDATION FAILED — %s. Pipeline ABORTED.", exc)
        return

    # ── Phase 4: Data Merge ───────────────────────────────────
    try:
        df_merged = phase4_merge(manifest)
    except Exception as exc:
        log.error("PHASE 4 FAILED — %s", exc, exc_info=True)
        return

    # ── Phase 5: Persistence ──────────────────────────────────
    try:
        new_rows = phase5_persist(df_merged)
    except Exception as exc:
        log.error("PHASE 5 FAILED — %s", exc, exc_info=True)
        return

    # ── Phase 6: Notification ─────────────────────────────────
    try:
        phase6_notify(df_merged)
    except Exception as exc:
        log.error("PHASE 6 FAILED — %s", exc, exc_info=True)
        # Notification failure should not lose data; log and continue
        return

    log.info("=" * 60)
    log.info("GlassOrchestrator pipeline completed successfully")
    log.info("=" * 60)


# ─── Entry Point ──────────────────────────────────────────────────────────────

if __name__ == "__main__":
    run_pipeline()
